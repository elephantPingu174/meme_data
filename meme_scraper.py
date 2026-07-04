import json
import logging
import os
import re
import time
from dataclasses import dataclass
from datetime import datetime, timezone
from typing import Any, Dict, List, Optional, Tuple
from urllib.parse import urljoin
from urllib.robotparser import RobotFileParser

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from requests import Session
from requests.exceptions import RequestException, Timeout


# ============================================================
# KONFIGURACJA
# ============================================================

TARGET_URL = "https://feargreedmeter.com/top-100-most-popular-meme-stocks-today"

# Zostawiam Playwright jako domyślny tryb, bo realnie odświeża stronę.
# Możesz zmienić na "requests", jeśli zwykły HTML wystarcza.
SCRAPER_MODE = "playwright"

OUTPUT_XLSX = "meme_stocks_snapshots.xlsx"
EXCEL_SHEET_NAME = "Snapshots"

LAST_SUCCESS_JSON = "last_successful_snapshot.json"

# Ranking aktualizuje się co 5 minut.
REFRESH_INTERVAL_SECONDS = 300
MIN_REFRESH_INTERVAL_SECONDS = 60

# Pobieranie będzie wyrównane do minut podzielnych przez 5:
# np. 23:10, 23:15, 23:20.
ALIGN_TO_5_MIN_BOUNDARY = True

# Mały bufor, żeby pobrać dane np. o 23:10:10, a nie idealnie o 23:10:00.
FETCH_AFTER_BOUNDARY_DELAY_SECONDS = 10

REQUEST_TIMEOUT_SECONDS = 20

USER_AGENT = (
    "EducationalMemeStockScraper/1.0 "
    "(low-frequency research project; contact: your-email@example.com)"
)

CHECK_ROBOTS_TXT = True
STOP_IF_ROBOTS_UNAVAILABLE = False

STOP_ON_403 = True
STOP_ON_429 = False

DEFAULT_429_SLEEP_SECONDS = 10 * 60
MAX_CONSECUTIVE_ERRORS = 5

FORCE_PAGE_REFRESH = True

# Domyślnie False, bo cache-buster typu ?ts=... może mocniej obciążać serwer/CDN.
USE_CACHE_BUSTER_QUERY_PARAM = False

# Jeśli snapshot jest identyczny jak poprzedni, nadal dopisuje go do Excela.
# Jeśli chcesz zapisywać tylko zmiany, ustaw False.
APPEND_UNCHANGED_SNAPSHOTS = True

EXCEL_HEADERS = [
    "data_pobrania",
    "kod",
    "nazwa",
    "upvotes",
    "mentions",
]


# ============================================================
# LOGOWANIE
# ============================================================

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)s | %(message)s",
)

logger = logging.getLogger("meme-stock-scraper")


# ============================================================
# WYJĄTKI
# ============================================================

class ScraperError(Exception):
    pass


class FatalScraperError(ScraperError):
    pass


class BlockedScraperError(FatalScraperError):
    pass


class NotFoundScraperError(FatalScraperError):
    pass


class RateLimitedScraperError(ScraperError):
    def __init__(self, message: str, retry_after_seconds: Optional[int] = None):
        super().__init__(message)
        self.retry_after_seconds = retry_after_seconds


class TransientScraperError(ScraperError):
    pass


class ParseScraperError(ScraperError):
    pass


# ============================================================
# MODEL DANYCH
# ============================================================

@dataclass
class MemeStockRow:
    fetched_at_utc: str
    rank: Optional[int]
    ticker: Optional[str]
    company_name: Optional[str]
    upvotes: Optional[int]
    mentions: Optional[int]
    mention_change: Optional[int]
    source_url: str
    raw_text: str

    def to_json_dict(self) -> Dict[str, Any]:
        return {
            "fetched_at_utc": self.fetched_at_utc,
            "rank": self.rank,
            "ticker": self.ticker,
            "company_name": self.company_name,
            "upvotes": self.upvotes,
            "mentions": self.mentions,
            "mention_change": self.mention_change,
            "source_url": self.source_url,
            "raw_text": self.raw_text,
        }


# ============================================================
# FUNKCJE POMOCNICZE
# ============================================================

def utc_now_iso() -> str:
    return datetime.now(timezone.utc).isoformat(timespec="seconds")


def clean_text(text: str) -> str:
    return re.sub(r"\s+", " ", text or "").strip()


def parse_int(value: Optional[str]) -> Optional[int]:
    if value is None:
        return None

    value = value.strip()

    if not value:
        return None

    value = value.replace(",", "")
    value = value.replace("$", "")
    value = value.replace("%", "")
    value = value.replace("−", "-")
    value = re.sub(r"\s+", "", value)

    try:
        return int(value)
    except ValueError:
        return None


def parse_retry_after_seconds(value: Optional[str]) -> Optional[int]:
    if not value:
        return None

    value = value.strip()

    if value.isdigit():
        return int(value)

    return None


def build_fetch_url(base_url: str) -> str:
    if not USE_CACHE_BUSTER_QUERY_PARAM:
        return base_url

    separator = "&" if "?" in base_url else "?"
    return f"{base_url}{separator}_scrape_ts={int(time.time())}"


def create_http_session() -> Session:
    session = requests.Session()

    session.headers.update({
        "User-Agent": USER_AGENT,
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Accept-Language": "en-US,en;q=0.8,pl;q=0.6",
        "Cache-Control": "no-cache",
        "Pragma": "no-cache",
    })

    return session


def compute_snapshot_fingerprint(rows: List[MemeStockRow]) -> str:
    compact = [
        {
            "rank": row.rank,
            "ticker": row.ticker,
            "company_name": row.company_name,
            "upvotes": row.upvotes,
            "mentions": row.mentions,
            "mention_change": row.mention_change,
        }
        for row in rows
    ]

    return json.dumps(compact, ensure_ascii=False, sort_keys=True)


def seconds_until_next_aligned_fetch() -> int:
    """
    Liczy czas do kolejnego pobrania wyrównanego do 5-minutowej granicy.

    Przykłady:
    - jeśli teraz jest 23:11:20, następne pobranie będzie około 23:15:10
    - jeśli teraz jest 23:15:12, następne pobranie będzie około 23:20:10

    Działa na czasie systemowym komputera.
    """
    if not ALIGN_TO_5_MIN_BOUNDARY:
        return max(REFRESH_INTERVAL_SECONDS, MIN_REFRESH_INTERVAL_SECONDS)

    now = time.time()

    interval = REFRESH_INTERVAL_SECONDS

    next_boundary = (int(now // interval) + 1) * interval
    next_fetch_time = next_boundary + FETCH_AFTER_BOUNDARY_DELAY_SECONDS

    sleep_seconds = int(round(next_fetch_time - now))

    return max(sleep_seconds, MIN_REFRESH_INTERVAL_SECONDS)


def sleep_until_next_scheduled_fetch() -> None:
    sleep_seconds = seconds_until_next_aligned_fetch()

    next_fetch_timestamp = time.time() + sleep_seconds
    next_fetch_local = datetime.fromtimestamp(next_fetch_timestamp).strftime("%Y-%m-%d %H:%M:%S")

    logger.info(
        "Czekam %s sekund. Następne pobranie około: %s",
        sleep_seconds,
        next_fetch_local,
    )

    time.sleep(sleep_seconds)


def check_robots_txt(session: Session, target_url: str) -> bool:
    robots_url = urljoin(target_url, "/robots.txt")

    logger.info("Sprawdzam robots.txt: %s", robots_url)

    try:
        response = session.get(
            robots_url,
            timeout=REQUEST_TIMEOUT_SECONDS,
        )
    except RequestException as exc:
        message = f"Nie udało się pobrać robots.txt: {exc}"

        if STOP_IF_ROBOTS_UNAVAILABLE:
            logger.error("%s. Przerywam.", message)
            return False

        logger.warning("%s. Kontynuuję ostrożnie.", message)
        return True

    if response.status_code == 404:
        message = "robots.txt nie istnieje albo zwrócił 404."

        if STOP_IF_ROBOTS_UNAVAILABLE:
            logger.error("%s Przerywam.", message)
            return False

        logger.warning("%s Kontynuuję ostrożnie.", message)
        return True

    if response.status_code != 200:
        message = f"robots.txt zwrócił HTTP {response.status_code}."

        if STOP_IF_ROBOTS_UNAVAILABLE:
            logger.error("%s Przerywam.", message)
            return False

        logger.warning("%s Kontynuuję ostrożnie.", message)
        return True

    parser = RobotFileParser()
    parser.set_url(robots_url)
    parser.parse(response.text.splitlines())

    allowed = parser.can_fetch(USER_AGENT, target_url)

    if not allowed:
        logger.error(
            "robots.txt nie pozwala temu User-Agentowi pobierać %s. Przerywam.",
            target_url,
        )
        return False

    logger.info("robots.txt pozwala na pobranie strony.")
    return True


# ============================================================
# POBIERANIE HTML: REQUESTS
# ============================================================

def fetch_html_requests(session: Session, url: str) -> str:
    fetch_url = build_fetch_url(url)

    logger.info("Pobieram stronę przez requests: %s", fetch_url)

    try:
        response = session.get(fetch_url, timeout=REQUEST_TIMEOUT_SECONDS)
    except Timeout as exc:
        raise TransientScraperError(f"Timeout połączenia: {exc}") from exc
    except RequestException as exc:
        raise TransientScraperError(f"Błąd połączenia: {exc}") from exc

    status = response.status_code

    if status == 200:
        return response.text

    if status == 403:
        raise BlockedScraperError(
            "HTTP 403: strona odmówiła dostępu. "
            "Nie obchodzę blokady. Rozważ oficjalne źródło danych albo kontakt z właścicielem strony."
        )

    if status == 404:
        raise NotFoundScraperError(
            "HTTP 404: strona nie została znaleziona. Sprawdź URL."
        )

    if status == 429:
        retry_after = parse_retry_after_seconds(response.headers.get("Retry-After"))

        raise RateLimitedScraperError(
            "HTTP 429: za dużo zapytań. Skrypt wydłuży przerwę albo zakończy działanie.",
            retry_after_seconds=retry_after,
        )

    if 500 <= status <= 599:
        raise TransientScraperError(f"HTTP {status}: błąd po stronie serwera.")

    if 400 <= status <= 499:
        raise FatalScraperError(
            f"HTTP {status}: błąd klienta. Nie ponawiam agresywnie."
        )

    raise TransientScraperError(f"Nieoczekiwany HTTP status: {status}")


# ============================================================
# POBIERANIE HTML: PLAYWRIGHT
# ============================================================

def fetch_html_playwright(page: Any, url: str, already_loaded: bool = False) -> str:
    fetch_url = build_fetch_url(url)

    if already_loaded and FORCE_PAGE_REFRESH and not USE_CACHE_BUSTER_QUERY_PARAM:
        logger.info("Odświeżam stronę przez Playwright: page.reload()")
    else:
        logger.info("Wchodzę na stronę przez Playwright: %s", fetch_url)

    try:
        if already_loaded and FORCE_PAGE_REFRESH and not USE_CACHE_BUSTER_QUERY_PARAM:
            response = page.reload(
                wait_until="domcontentloaded",
                timeout=REQUEST_TIMEOUT_SECONDS * 1000,
            )
        else:
            response = page.goto(
                fetch_url,
                wait_until="domcontentloaded",
                timeout=REQUEST_TIMEOUT_SECONDS * 1000,
            )

        if response is None:
            raise TransientScraperError("Playwright nie zwrócił odpowiedzi HTTP.")

        status = response.status

        if status == 403:
            raise BlockedScraperError(
                "HTTP 403: strona odmówiła dostępu. "
                "Nie obchodzę blokady. Rozważ oficjalne źródło danych."
            )

        if status == 404:
            raise NotFoundScraperError(
                "HTTP 404: strona nie została znaleziona. Sprawdź URL."
            )

        if status == 429:
            retry_after = parse_retry_after_seconds(response.headers.get("retry-after"))

            raise RateLimitedScraperError(
                "HTTP 429: za dużo zapytań.",
                retry_after_seconds=retry_after,
            )

        if 500 <= status <= 599:
            raise TransientScraperError(f"HTTP {status}: błąd po stronie serwera.")

        if 400 <= status <= 499:
            raise FatalScraperError(f"HTTP {status}: błąd klienta.")

        try:
            page.wait_for_load_state("networkidle", timeout=10_000)
        except Exception:
            logger.warning("Nie osiągnięto networkidle, próbuję parsować aktualny HTML.")

        return page.content()

    except ScraperError:
        raise
    except Exception as exc:
        raise TransientScraperError(f"Błąd Playwright: {exc}") from exc


# ============================================================
# PARSOWANIE HTML
# ============================================================

def parse_html(html: str, fetched_at_utc: str) -> List[MemeStockRow]:
    soup = BeautifulSoup(html, "lxml")

    rows = parse_table_rows(soup, fetched_at_utc)

    if rows:
        logger.info("Znaleziono %s rekordów metodą tabel HTML.", len(rows))
        return rows

    rows = parse_card_or_link_rows(soup, fetched_at_utc)

    if rows:
        logger.info("Znaleziono %s rekordów metodą kart/linków.", len(rows))
        return rows

    raise ParseScraperError(
        "Nie znaleziono rekordów rankingu. "
        "Możliwe, że strona zmieniła HTML albo ładuje dane JavaScriptem z osobnego endpointu."
    )


def parse_table_rows(soup: BeautifulSoup, fetched_at_utc: str) -> List[MemeStockRow]:
    parsed_rows: List[MemeStockRow] = []

    for table in soup.select("table"):
        headers = [
            clean_text(th.get_text(" ", strip=True)).lower()
            for th in table.select("thead th")
        ]

        if not headers:
            first_row = table.select_one("tr")

            if first_row:
                headers = [
                    clean_text(cell.get_text(" ", strip=True)).lower()
                    for cell in first_row.select("th,td")
                ]

        header_blob = " ".join(headers)

        if not any(key in header_blob for key in ["ticker", "symbol", "mentions", "upvotes"]):
            continue

        body_rows = table.select("tbody tr") or table.select("tr")[1:]

        for tr in body_rows:
            cells = [
                clean_text(td.get_text(" ", strip=True))
                for td in tr.select("td")
            ]

            if len(cells) < 2:
                continue

            row_dict = map_table_cells(headers, cells)
            raw_text = clean_text(tr.get_text(" ", strip=True))

            if not row_dict.get("ticker"):
                fallback = parse_ranked_text_row(raw_text, fetched_at_utc)

                if fallback:
                    parsed_rows.append(fallback)

                continue

            parsed_rows.append(
                MemeStockRow(
                    fetched_at_utc=fetched_at_utc,
                    rank=row_dict.get("rank"),
                    ticker=row_dict.get("ticker"),
                    company_name=row_dict.get("company_name"),
                    upvotes=row_dict.get("upvotes"),
                    mentions=row_dict.get("mentions"),
                    mention_change=row_dict.get("mention_change"),
                    source_url=TARGET_URL,
                    raw_text=raw_text,
                )
            )

    return deduplicate_rows(parsed_rows)


def map_table_cells(headers: List[str], cells: List[str]) -> Dict[str, Any]:
    result: Dict[str, Any] = {
        "rank": None,
        "ticker": None,
        "company_name": None,
        "upvotes": None,
        "mentions": None,
        "mention_change": None,
    }

    for idx, value in enumerate(cells):
        header = headers[idx] if idx < len(headers) else ""

        if any(key in header for key in ["rank", "#", "position", "pozycja"]):
            result["rank"] = parse_int(value)

        elif any(key in header for key in ["ticker", "symbol", "kod"]):
            result["ticker"] = value.upper() if value else None

        elif any(key in header for key in ["company", "name", "spółka", "nazwa"]):
            result["company_name"] = value or None

        elif "upvote" in header:
            result["upvotes"] = parse_int(value)

        elif "mention" in header and "change" not in header:
            result["mentions"] = parse_int(value)

        elif "change" in header or "delta" in header or "trend" in header:
            result["mention_change"] = parse_int(value)

    return result


def parse_card_or_link_rows(soup: BeautifulSoup, fetched_at_utc: str) -> List[MemeStockRow]:
    candidates: List[str] = []

    for element in soup.select("a"):
        text = clean_text(element.get_text(" ", strip=True))

        if looks_like_rank_row(text):
            candidates.append(text)

    if not candidates:
        selectors = [
            "[class*='stock']",
            "[class*='rank']",
            "[class*='ticker']",
            "[class*='card']",
            "[class*='item']",
        ]

        for selector in selectors:
            for element in soup.select(selector):
                text = clean_text(element.get_text(" ", strip=True))

                if looks_like_rank_row(text):
                    candidates.append(text)

    parsed_rows: List[MemeStockRow] = []

    for text in candidates:
        row = parse_ranked_text_row(text, fetched_at_utc)

        if row:
            parsed_rows.append(row)

    return deduplicate_rows(parsed_rows)


def looks_like_rank_row(text: str) -> bool:
    if not text:
        return False

    return bool(
        re.match(
            r"^#\s*\d{1,3}\s*\.?\s+[A-Z][A-Z0-9.\-]{0,9}\b",
            text,
        )
    )


def parse_ranked_text_row(text: str, fetched_at_utc: str) -> Optional[MemeStockRow]:
    """
    Parser dla formatu podobnego do:

        # 1 . MU Micron Technology 2,893 677 + 575

    Wynik:
        rank = 1
        ticker = MU
        company_name = Micron Technology
        upvotes = 2893
        mentions = 677
        mention_change = 575
    """
    pattern = re.compile(
        r"^#\s*(?P<rank>\d{1,3})\s*\.?\s+"
        r"(?P<ticker>[A-Z][A-Z0-9.\-]{0,9})\s+"
        r"(?P<rest>.+)$"
    )

    match = pattern.match(text)

    if not match:
        return None

    rank = parse_int(match.group("rank"))
    ticker = match.group("ticker").strip().upper()
    rest = clean_text(match.group("rest"))

    tokens = rest.split()

    if len(tokens) < 3:
        return MemeStockRow(
            fetched_at_utc=fetched_at_utc,
            rank=rank,
            ticker=ticker,
            company_name=rest or None,
            upvotes=None,
            mentions=None,
            mention_change=None,
            source_url=TARGET_URL,
            raw_text=text,
        )

    mention_change, tokens = pop_trailing_signed_int(tokens)

    mentions = None
    upvotes = None

    if len(tokens) >= 1:
        mentions_candidate = parse_int(tokens[-1])

        if mentions_candidate is not None:
            mentions = mentions_candidate
            tokens = tokens[:-1]

    if len(tokens) >= 1:
        upvotes_candidate = parse_int(tokens[-1])

        if upvotes_candidate is not None:
            upvotes = upvotes_candidate
            tokens = tokens[:-1]

    company_name = " ".join(tokens).strip() or None

    return MemeStockRow(
        fetched_at_utc=fetched_at_utc,
        rank=rank,
        ticker=ticker,
        company_name=company_name,
        upvotes=upvotes,
        mentions=mentions,
        mention_change=mention_change,
        source_url=TARGET_URL,
        raw_text=text,
    )


def pop_trailing_signed_int(tokens: List[str]) -> Tuple[Optional[int], List[str]]:
    if not tokens:
        return None, tokens

    if len(tokens) >= 2 and tokens[-2] in {"+", "-"}:
        number = parse_int(tokens[-1])

        if number is not None:
            sign = 1 if tokens[-2] == "+" else -1
            return sign * abs(number), tokens[:-2]

    last = tokens[-1].replace("−", "-")

    if re.match(r"^[+-]\d[\d,]*$", last):
        return parse_int(last), tokens[:-1]

    return None, tokens


def deduplicate_rows(rows: List[MemeStockRow]) -> List[MemeStockRow]:
    seen = set()
    unique_rows: List[MemeStockRow] = []

    for row in rows:
        key = (row.rank, row.ticker)

        if key in seen:
            continue

        seen.add(key)
        unique_rows.append(row)

    unique_rows.sort(key=lambda r: r.rank if r.rank is not None else 9999)

    return unique_rows


# ============================================================
# ZAPIS DO EXCELA
# ============================================================

def append_rows_to_excel(rows: List[MemeStockRow], path: str) -> None:
    """
    Dopisuje dane do pliku Excel.

    Kolumny:
        data_pobrania | kod | nazwa | upvotes | mentions
    """
    if os.path.exists(path):
        workbook = load_workbook(path)

        if EXCEL_SHEET_NAME in workbook.sheetnames:
            sheet = workbook[EXCEL_SHEET_NAME]
        else:
            sheet = workbook.create_sheet(EXCEL_SHEET_NAME)
            sheet.append(EXCEL_HEADERS)

    else:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = EXCEL_SHEET_NAME
        sheet.append(EXCEL_HEADERS)

    for row in rows:
        sheet.append([
            row.fetched_at_utc,
            row.ticker,
            row.company_name,
            row.upvotes,
            row.mentions,
        ])

    sheet.column_dimensions["A"].width = 28
    sheet.column_dimensions["B"].width = 14
    sheet.column_dimensions["C"].width = 40
    sheet.column_dimensions["D"].width = 14
    sheet.column_dimensions["E"].width = 14

    try:
        temp_path = f"{path}.tmp.xlsx"
        workbook.save(temp_path)
        os.replace(temp_path, path)

        logger.info("Dopisano %s rekordów do Excela: %s", len(rows), path)

    except PermissionError:
        logger.error(
            "Nie udało się zapisać pliku Excel. "
            "Sprawdź, czy plik nie jest aktualnie otwarty w Excelu: %s",
            path,
        )

    finally:
        workbook.close()


def save_last_successful_snapshot(rows: List[MemeStockRow], path: str) -> None:
    payload = {
        "saved_at_utc": utc_now_iso(),
        "source_url": TARGET_URL,
        "rows_count": len(rows),
        "rows": [row.to_json_dict() for row in rows],
    }

    temp_path = f"{path}.tmp"

    with open(temp_path, mode="w", encoding="utf-8") as file:
        json.dump(payload, file, ensure_ascii=False, indent=2)

    os.replace(temp_path, path)

    logger.info("Zapisano ostatni poprawny snapshot: %s", path)


# ============================================================
# BACKOFF DLA BŁĘDÓW
# ============================================================

def compute_error_sleep_seconds(consecutive_errors: int) -> int:
    base = max(REFRESH_INTERVAL_SECONDS, MIN_REFRESH_INTERVAL_SECONDS)

    sleep_seconds = base * (2 ** max(0, consecutive_errors - 1))

    return min(sleep_seconds, 15 * 60)


def sleep_safely(seconds: int) -> None:
    seconds = max(seconds, MIN_REFRESH_INTERVAL_SECONDS)

    logger.info("Czekam %s sekund przed kolejną próbą.", seconds)
    time.sleep(seconds)


# ============================================================
# GŁÓWNA PĘTLA: REQUESTS
# ============================================================

def run_requests_scraper() -> None:
    session = create_http_session()

    if CHECK_ROBOTS_TXT:
        allowed = check_robots_txt(session, TARGET_URL)

        if not allowed:
            return

    consecutive_errors = 0
    last_fingerprint: Optional[str] = None

    while True:
        try:
            fetched_at = utc_now_iso()

            html = fetch_html_requests(session, TARGET_URL)
            rows = parse_html(html, fetched_at)

            if not rows:
                raise ParseScraperError("Parser zwrócił 0 rekordów.")

            if len(rows) < 50:
                logger.warning(
                    "Znaleziono tylko %s rekordów. "
                    "To może oznaczać zmianę HTML albo częściowe ładowanie JS.",
                    len(rows),
                )

            current_fingerprint = compute_snapshot_fingerprint(rows)

            if last_fingerprint is not None and current_fingerprint == last_fingerprint:
                logger.info("Snapshot jest identyczny jak poprzedni.")

                if not APPEND_UNCHANGED_SNAPSHOTS:
                    logger.info("Nie dopisuję identycznego snapshotu do Excela.")
                    sleep_until_next_scheduled_fetch()
                    continue

            else:
                logger.info("Snapshot różni się od poprzedniego albo jest pierwszym pobraniem.")

            append_rows_to_excel(rows, OUTPUT_XLSX)
            save_last_successful_snapshot(rows, LAST_SUCCESS_JSON)

            last_fingerprint = current_fingerprint
            consecutive_errors = 0

            sleep_until_next_scheduled_fetch()

        except RateLimitedScraperError as exc:
            consecutive_errors += 1

            logger.warning("%s", exc)

            if STOP_ON_429:
                logger.error("STOP_ON_429=True, kończę działanie.")
                break

            if consecutive_errors >= MAX_CONSECUTIVE_ERRORS:
                logger.error(
                    "Za dużo błędów 429 lub innych błędów z rzędu. "
                    "Kończę działanie. Rozważ oficjalne API albo kontakt z właścicielem strony."
                )
                break

            sleep_seconds = exc.retry_after_seconds or DEFAULT_429_SLEEP_SECONDS
            sleep_safely(sleep_seconds)

        except BlockedScraperError as exc:
            logger.error("%s", exc)

            if STOP_ON_403:
                logger.error(
                    "Kończę działanie, żeby nie próbować obchodzić blokady. "
                    "Rozważ oficjalne źródło danych."
                )
                break

            consecutive_errors += 1
            sleep_safely(compute_error_sleep_seconds(consecutive_errors))

        except FatalScraperError as exc:
            logger.error("%s", exc)
            break

        except (TransientScraperError, ParseScraperError) as exc:
            consecutive_errors += 1

            logger.warning("Błąd chwilowy/parsingowy: %s", exc)

            if consecutive_errors >= MAX_CONSECUTIVE_ERRORS:
                logger.error(
                    "Osiągnięto limit błędów z rzędu. "
                    "Kończę działanie, żeby nie wykonywać agresywnych prób."
                )
                break

            sleep_safely(compute_error_sleep_seconds(consecutive_errors))

        except KeyboardInterrupt:
            logger.info("Zatrzymano ręcznie przez Ctrl+C.")
            break


# ============================================================
# GŁÓWNA PĘTLA: PLAYWRIGHT
# ============================================================

def run_playwright_scraper() -> None:
    session = create_http_session()

    if CHECK_ROBOTS_TXT:
        allowed = check_robots_txt(session, TARGET_URL)

        if not allowed:
            return

    try:
        from playwright.sync_api import sync_playwright
    except ImportError as exc:
        raise RuntimeError(
            "Brakuje Playwright. Zainstaluj: "
            "pip install playwright && playwright install chromium"
        ) from exc

    consecutive_errors = 0
    last_fingerprint: Optional[str] = None
    page_already_loaded = False

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)

        context = browser.new_context(
            user_agent=USER_AGENT,
            locale="en-US",
        )

        context.set_extra_http_headers({
            "Cache-Control": "no-cache",
            "Pragma": "no-cache",
        })

        page = context.new_page()

        try:
            while True:
                try:
                    fetched_at = utc_now_iso()

                    html = fetch_html_playwright(
                        page=page,
                        url=TARGET_URL,
                        already_loaded=page_already_loaded,
                    )

                    page_already_loaded = True

                    rows = parse_html(html, fetched_at)

                    if not rows:
                        raise ParseScraperError("Parser zwrócił 0 rekordów.")

                    if len(rows) < 50:
                        logger.warning(
                            "Znaleziono tylko %s rekordów. "
                            "Sprawdź selektory HTML.",
                            len(rows),
                        )

                    current_fingerprint = compute_snapshot_fingerprint(rows)

                    if last_fingerprint is not None and current_fingerprint == last_fingerprint:
                        logger.info("Snapshot jest identyczny jak poprzedni.")

                        if not APPEND_UNCHANGED_SNAPSHOTS:
                            logger.info("Nie dopisuję identycznego snapshotu do Excela.")
                            sleep_until_next_scheduled_fetch()
                            continue

                    else:
                        logger.info("Snapshot różni się od poprzedniego albo jest pierwszym pobraniem.")

                    append_rows_to_excel(rows, OUTPUT_XLSX)
                    save_last_successful_snapshot(rows, LAST_SUCCESS_JSON)

                    last_fingerprint = current_fingerprint
                    consecutive_errors = 0

                    sleep_until_next_scheduled_fetch()

                except RateLimitedScraperError as exc:
                    consecutive_errors += 1

                    logger.warning("%s", exc)

                    if STOP_ON_429:
                        logger.error("STOP_ON_429=True, kończę działanie.")
                        break

                    if consecutive_errors >= MAX_CONSECUTIVE_ERRORS:
                        logger.error(
                            "Za dużo błędów 429 lub innych błędów z rzędu. "
                            "Kończę działanie."
                        )
                        break

                    sleep_seconds = exc.retry_after_seconds or DEFAULT_429_SLEEP_SECONDS
                    sleep_safely(sleep_seconds)

                except BlockedScraperError as exc:
                    logger.error("%s", exc)

                    if STOP_ON_403:
                        logger.error(
                            "Kończę działanie, żeby nie obchodzić blokady. "
                            "Rozważ oficjalne źródło danych."
                        )
                        break

                    consecutive_errors += 1
                    sleep_safely(compute_error_sleep_seconds(consecutive_errors))

                except FatalScraperError as exc:
                    logger.error("%s", exc)
                    break

                except (TransientScraperError, ParseScraperError) as exc:
                    consecutive_errors += 1

                    logger.warning("Błąd chwilowy/parsingowy: %s", exc)

                    if consecutive_errors >= MAX_CONSECUTIVE_ERRORS:
                        logger.error(
                            "Osiągnięto limit błędów z rzędu. "
                            "Kończę działanie."
                        )
                        break

                    sleep_safely(compute_error_sleep_seconds(consecutive_errors))

                except KeyboardInterrupt:
                    logger.info("Zatrzymano ręcznie przez Ctrl+C.")
                    break

        finally:
            context.close()
            browser.close()


# ============================================================
# MAIN
# ============================================================

def main() -> None:
    if REFRESH_INTERVAL_SECONDS < MIN_REFRESH_INTERVAL_SECONDS:
        raise ValueError(
            f"REFRESH_INTERVAL_SECONDS nie może być mniejsze niż "
            f"{MIN_REFRESH_INTERVAL_SECONDS}."
        )

    logger.info("Start scrapera.")
    logger.info("Tryb: %s", SCRAPER_MODE)
    logger.info("Interwał: %s sekund", REFRESH_INTERVAL_SECONDS)
    logger.info("Wyrównanie do 5-minutowych granic: %s", ALIGN_TO_5_MIN_BOUNDARY)
    logger.info("Bufor po granicy: %s sekund", FETCH_AFTER_BOUNDARY_DELAY_SECONDS)
    logger.info("URL: %s", TARGET_URL)
    logger.info("Plik Excel: %s", OUTPUT_XLSX)
    logger.info("Force refresh: %s", FORCE_PAGE_REFRESH)
    logger.info("Cache buster query param: %s", USE_CACHE_BUSTER_QUERY_PARAM)

    if SCRAPER_MODE == "requests":
        run_requests_scraper()

    elif SCRAPER_MODE == "playwright":
        run_playwright_scraper()

    else:
        raise ValueError('SCRAPER_MODE musi mieć wartość "requests" albo "playwright".')

    logger.info("Scraper zakończył działanie.")


if __name__ == "__main__":
    main()
